import openpyxl

class readWriteExcel:
    wb = None
    excelFilePath = None

    def __init__(self,FilePath):
        global wb
        global excelFilePath
        excelFilePath = FilePath
        wb = openpyxl.load_workbook(excelFilePath)

    def readByColIndex(self,sheetName,rowIndex,colIndex):
        sheet = wb[sheetName]
        return sheet.cell(rowIndex,colIndex).value

    def readByColName(self,sheetName,rowIndex,colName):
        sheet = wb[sheetName]
        colIndex = 1
        while(sheet.cell(row=1, column=colIndex).value != ''):
            if(colName==sheet.cell(row=1, column=colIndex).value):
                break
            colIndex = colIndex+1
            return sheet.cell(rowIndex,colIndex).value